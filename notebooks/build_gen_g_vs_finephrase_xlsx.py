"""Build an Excel workbook comparing our Gen G rephrase outputs against the
FinePhrase HuggingFace dataset (HuggingFaceFW/finephrase), side-by-side per format.

Tabs:
  tutorial / faq / math   -> FinePhrase + Ours (5-column layout)
  table                   -> FinePhrase only (we don't produce table); 3-col layout
  narrative / explanation -> Ours only (not released on HF); 3-col layout

Output: D:/hist_LLM/periods/1900_1949/posttraining_data/synthetic/gen_g_vs_finephrase_comparison.xlsx
"""
from __future__ import annotations

import json
import random
from pathlib import Path

from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
GEN_G_DIR = Path("D:/hist_LLM/periods/1900_1949/posttraining_data/synthetic/by_generator")
OUTPUT_XLSX = Path(
    "D:/hist_LLM/periods/1900_1949/posttraining_data/synthetic/gen_g_vs_finephrase_comparison.xlsx"
)

N_SAMPLES = 8
HF_REPO = "HuggingFaceFW/finephrase"
HF_FORMATS = ["faq", "math", "table", "tutorial"]
OUR_FORMATS = ["tutorial", "faq", "math", "narrative", "explanation"]

# Sheet order / flags:  (sheet_name, has_finephrase, has_ours)
SHEETS = [
    ("tutorial",    True,  True),
    ("faq",         True,  True),
    ("math",        True,  True),
    ("table",       True,  False),  # FinePhrase-only
    ("narrative",   False, True),   # Ours-only
    ("explanation", False, True),   # Ours-only
]

# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #
BANNER_FILL = PatternFill("solid", fgColor="1F4E79")
BANNER_FONT = Font(bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


# --------------------------------------------------------------------------- #
# Data loaders
# --------------------------------------------------------------------------- #
def load_finephrase_samples(config: str, n: int, seed: int = 42) -> list[dict]:
    """Fetch ~n random samples (source + rephrased) from the HF dataset."""
    print(f"[HF] loading {HF_REPO} :: {config}  (streaming)", flush=True)
    # Reservoir sample from the streaming dataset so we don't download all shards.
    ds = load_dataset(HF_REPO, config, split="train", streaming=True)
    rng = random.Random(seed)
    reservoir: list[dict] = []
    pool_cap = 200  # take first 200 then sample n
    for i, row in enumerate(ds):
        if i >= pool_cap:
            break
        reservoir.append(row)
    rng.shuffle(reservoir)
    picks = reservoir[:n]
    out = []
    for r in picks:
        src = r.get("text", "") or ""
        rr = r.get("rollout_results") or []
        rephrased = rr[0].get("text", "") if rr else ""
        out.append({"source": src, "output": rephrased})
    print(f"[HF] {config}: returned {len(out)} samples", flush=True)
    return out


def load_our_samples(fmt: str, n: int, seed: int = 42) -> list[dict]:
    """Load up to n conversations from our gen_g JSONL for a given format."""
    path = GEN_G_DIR / f"gen_g_rephrase_rephrase_{fmt}.jsonl"
    rows: list[dict] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    rng = random.Random(seed)
    rng.shuffle(rows)
    picks = rows[:n]
    out = []
    for c in picks:
        msgs = c.get("messages", [])
        user = next((m["content"] for m in msgs if m.get("role") == "user"), "")
        asst = next((m["content"] for m in msgs if m.get("role") == "assistant"), "")
        tag = f"{c.get('doc_name', '?')} (chunk {c.get('chunk_idx', '?')})"
        out.append({"source_tag": tag, "source": user, "output": asst})
    print(f"[ours] {fmt}: returned {len(out)} samples (of {len(rows)} total rows)", flush=True)
    return out


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def _write_banner(ws, text: str, n_cols: int) -> None:
    ws.cell(row=1, column=1, value=text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1)
    c.fill = BANNER_FILL
    c.font = BANNER_FONT
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28


def _write_headers(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[2].height = 22


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_cell(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value if value is not None else "")
    c.alignment = WRAP


def build_sheet_both(ws, fmt: str, fp_rows: list[dict], our_rows: list[dict]) -> int:
    headers = [
        "Index",
        "FinePhrase Source",
        "FinePhrase Output",
        "Our Source (doc_name/chunk)",
        "Our Output",
    ]
    _write_banner(ws, f"FORMAT: {fmt} - FinePhrase HF vs Ours", len(headers))
    _write_headers(ws, headers)
    _set_col_widths(ws, [10, 60, 60, 60, 60])

    n = max(len(fp_rows), len(our_rows))
    for i in range(n):
        r = 3 + i
        _write_cell(ws, r, 1, str(i + 1))
        if i < len(fp_rows):
            _write_cell(ws, r, 2, fp_rows[i]["source"])
            _write_cell(ws, r, 3, fp_rows[i]["output"])
        if i < len(our_rows):
            _write_cell(ws, r, 4, our_rows[i]["source_tag"] + "\n\n" + our_rows[i]["source"])
            _write_cell(ws, r, 5, our_rows[i]["output"])
    return n


def build_sheet_fp_only(ws, fmt: str, fp_rows: list[dict]) -> int:
    headers = ["Index", "Source", "Output"]
    _write_banner(ws, f"FORMAT: {fmt} - FinePhrase HF only (N/A - not in our pipeline)", len(headers))
    _write_headers(ws, headers)
    _set_col_widths(ws, [10, 60, 60])
    for i, row in enumerate(fp_rows):
        r = 3 + i
        _write_cell(ws, r, 1, str(i + 1))
        _write_cell(ws, r, 2, row["source"])
        _write_cell(ws, r, 3, row["output"])
    # Annotation row noting the N/A status on the ours side (already in banner).
    return len(fp_rows)


def build_sheet_ours_only(ws, fmt: str, our_rows: list[dict]) -> int:
    headers = ["Index", "Our Source (doc_name/chunk)", "Our Output"]
    _write_banner(
        ws,
        f"FORMAT: {fmt} - Ours only (not released in HF dataset)",
        len(headers),
    )
    _write_headers(ws, headers)
    _set_col_widths(ws, [10, 60, 60])
    for i, row in enumerate(our_rows):
        r = 3 + i
        _write_cell(ws, r, 1, str(i + 1))
        _write_cell(ws, r, 2, row["source_tag"] + "\n\n" + row["source"])
        _write_cell(ws, r, 3, row["output"])
    return len(our_rows)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    # Pre-fetch data.
    fp_data = {fmt: load_finephrase_samples(fmt, N_SAMPLES) for fmt in HF_FORMATS}
    our_data = {fmt: load_our_samples(fmt, N_SAMPLES) for fmt in OUR_FORMATS}

    wb = Workbook()
    # Drop default sheet.
    default = wb.active
    wb.remove(default)

    tab_rows: dict[str, int] = {}
    for name, has_fp, has_ours in SHEETS:
        ws = wb.create_sheet(title=name)
        if has_fp and has_ours:
            n = build_sheet_both(ws, name, fp_data[name], our_data[name])
        elif has_fp and not has_ours:
            n = build_sheet_fp_only(ws, name, fp_data[name])
        elif has_ours and not has_fp:
            n = build_sheet_ours_only(ws, name, our_data[name])
        else:
            n = 0
        ws.freeze_panes = "A3"
        tab_rows[name] = n

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"[OK] wrote {OUTPUT_XLSX}", flush=True)

    # ------------------------------------------------------------------ #
    # Verify by re-opening.
    # ------------------------------------------------------------------ #
    from openpyxl import load_workbook
    wb2 = load_workbook(OUTPUT_XLSX)
    print(f"[verify] tabs: {wb2.sheetnames}", flush=True)
    for name in wb2.sheetnames:
        ws = wb2[name]
        # Count data rows (row 3 onward with non-empty col 1).
        data_rows = 0
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=1).value not in (None, ""):
                data_rows += 1
        print(f"[verify] {name}: {data_rows} data rows (banner+header+rows max={ws.max_row})",
              flush=True)

    # Spot check tutorial tab.
    ws = wb2["tutorial"]
    fp_out = (ws.cell(row=3, column=3).value or "")[:80]
    our_out = (ws.cell(row=3, column=5).value or "")[:80]
    print(f"[spot] tutorial row1 FinePhrase output[:80]: {fp_out!r}", flush=True)
    print(f"[spot] tutorial row1 Ours output[:80]:       {our_out!r}", flush=True)


if __name__ == "__main__":
    main()
