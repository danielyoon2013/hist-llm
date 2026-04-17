"""Build a single Excel workbook with each external benchmark on its own tab.

Each tab compares our training data (Gen A-F samples) against the benchmark's
WRONGLY-PREDICTED test items (so we can study failure modes).

Output: D:/hist_LLM/periods/1900_1949/error_analysis_new_prompt/training_vs_benchmark.xlsx
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------- config -----------------------------

PERIOD_DIR = Path("D:/hist_LLM/periods/1900_1949")
GEN_DIR = PERIOD_DIR / "posttraining_data" / "synthetic" / "by_generator"
BENCH_DIR = PERIOD_DIR / "error_analysis_new_prompt" / "sft_final"
OUT_PATH = PERIOD_DIR / "error_analysis_new_prompt" / "training_vs_benchmark.xlsx"

SAMPLE_N = 50

# benchmark name -> (gen file, gen letter, training banner, benchmark banner, details file)
BENCHMARKS = [
    (
        "ARC-Challenge",
        "gen_a_factual_mc4.jsonl",
        "A",
        "GENERATOR A \u2014 Factual QA / Science (training data shown below)",
        "EXTERNAL BENCHMARK: ARC-Challenge \u2014 wrong predictions only (tested against Generator A above)",
        "ARC-Challenge_details.jsonl",
    ),
    (
        "HellaSwag",
        "gen_e_completion_mc4.jsonl",
        "E",
        "GENERATOR E \u2014 Scene Completion / How-To (training data shown below)",
        "EXTERNAL BENCHMARK: HellaSwag \u2014 wrong predictions only (tested against Generator E above)",
        "HellaSwag_details.jsonl",
    ),
    (
        "RACE-Middle",
        "gen_c_comprehension_mc4_passage.jsonl",
        "C",
        "GENERATOR C \u2014 Reading Comprehension (training data shown below)",
        "EXTERNAL BENCHMARK: RACE-Middle \u2014 wrong predictions only (tested against Generator C above)",
        "RACE-Middle_details.jsonl",
    ),
    (
        "RACE-High",
        "gen_c_comprehension_mc4_passage.jsonl",
        "C",
        "GENERATOR C \u2014 Reading Comprehension (training data shown below)",
        "EXTERNAL BENCHMARK: RACE-High \u2014 wrong predictions only (tested against Generator C above)",
        "RACE-High_details.jsonl",
    ),
    (
        "Winogrande",
        "gen_f_instruct_mc2.jsonl",
        "F",
        "GENERATOR F \u2014 Pronoun / Commonsense Reference (training data shown below)",
        "EXTERNAL BENCHMARK: Winogrande \u2014 wrong predictions only (tested against Generator F above)",
        "Winogrande_details.jsonl",
    ),
    (
        "PIQA",
        "gen_b_cot_mc2.jsonl",
        "B",
        "GENERATOR B \u2014 Physical Commonsense (training data shown below)",
        "EXTERNAL BENCHMARK: PIQA \u2014 wrong predictions only (tested against Generator B above)",
        "PIQA_details.jsonl",
    ),
    (
        "GSM-MC",
        "gen_d_quantitative_mc4.jsonl",
        "D",
        "GENERATOR D \u2014 Quantitative / Math Word Problems (training data shown below)",
        "EXTERNAL BENCHMARK: GSM-MC \u2014 wrong predictions only (tested against Generator D above)",
        "GSM-MC_details.jsonl",
    ),
]


# ----------------------------- parsing -----------------------------

# Lines look like: "- choice text=A"  (choice text may contain '=', so split on
# the LAST '=' followed by a single capital letter at end of line.)
_CHOICE_RE = re.compile(r"^-\s+(.*?)=([A-D])\s*$")


def parse_question_block(question_text: str) -> tuple[str, dict[str, str]]:
    """Return (stem, {letter: choice_text}) parsed from a formatted MC question.

    The format used everywhere in this project:

        Multiple Choice question: <stem...possibly multi-line, possibly w/ Passage:>
        - choice text=A
        - choice text=B
        ...
        Respond only with the letter of the correct answer.
    """
    # strip the trailing instruction line if present
    text = question_text.rstrip()
    text = re.sub(
        r"\n\s*Respond only with the letter of the correct answer\.\s*$",
        "",
        text,
    )

    # strip leading "Multiple Choice question:" tag
    text = re.sub(r"^\s*Multiple Choice question:\s*", "", text)

    lines = text.split("\n")

    # Walk from the bottom collecting choice lines.
    choices: dict[str, str] = {}
    cut_idx = len(lines)
    for i in range(len(lines) - 1, -1, -1):
        line = lines[i].rstrip()
        if not line.strip():
            continue
        m = _CHOICE_RE.match(line.strip())
        if m:
            choices[m.group(2)] = m.group(1).strip()
            cut_idx = i
        else:
            # first non-choice non-blank line from the bottom -> stem ends here
            break

    stem_lines = lines[:cut_idx]
    stem = "\n".join(stem_lines).strip()
    return stem, choices


def gen_sample_to_row(idx: int, conv: dict) -> list:
    """Convert a generator JSONL conversation to a 7-col training row."""
    user_msg = next(m["content"] for m in conv["messages"] if m["role"] == "user")
    answer = next(m["content"] for m in conv["messages"] if m["role"] == "assistant").strip()
    stem, choices = parse_question_block(user_msg)
    return [
        idx,
        stem,
        choices.get("A", ""),
        choices.get("B", ""),
        choices.get("C", ""),
        choices.get("D", ""),
        answer,
    ]


def benchmark_sample_to_row(idx: int, item: dict) -> list:
    """Convert a benchmark details JSONL item to a 9-col benchmark row.

    Adds Predicted (letter) + Confidence (percent) at the end.
    """
    stem, choices = parse_question_block(item["question"])
    confidence = item.get("confidence")
    if isinstance(confidence, (int, float)):
        conf_str = f"{round(confidence * 100)}%"
    else:
        conf_str = ""
    return [
        idx,
        stem,
        choices.get("A", ""),
        choices.get("B", ""),
        choices.get("C", ""),
        choices.get("D", ""),
        item.get("expected", ""),
        item.get("predicted", ""),
        conf_str,
    ]


def read_jsonl(path: Path, limit: int | None = None) -> list[dict]:
    out: list[dict] = []
    with path.open("r", encoding="utf-8") as fh:
        for i, line in enumerate(fh):
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                continue
            if limit is not None and len(out) >= limit:
                break
    return out


# ----------------------------- workbook -----------------------------

TRAIN_HEADER_COLS = ["Index", "Question", "Choice_A", "Choice_B", "Choice_C", "Choice_D", "Correct"]
BENCH_HEADER_COLS = TRAIN_HEADER_COLS + ["Predicted", "Confidence"]
# Index=12, Question=60, Choice_*=30, Correct/Predicted/Confidence=12
COL_WIDTHS = [12, 60, 30, 30, 30, 30, 12, 12, 12]

TRAIN_SECTION_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
BENCH_SECTION_FILL = PatternFill(start_color="FFB22222", end_color="FFB22222", fill_type="solid")
COL_HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

SECTION_FONT = Font(bold=True, size=14, color="FFFFFFFF")
SECTION_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
COL_HEADER_FONT = Font(bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

SECTION_ROW_HEIGHT = 28


def write_section(
    ws,
    title: str,
    header_cols: list[str],
    rows: list[list],
    start_row: int,
    section_fill: PatternFill,
) -> int:
    """Write a section header, column header, and data rows. Return next free row."""
    n_cols = len(header_cols)
    # section title spanning all section cols, styled as a banner
    section_cell = ws.cell(row=start_row, column=1, value=title)
    section_cell.font = SECTION_FONT
    section_cell.fill = section_fill
    section_cell.alignment = SECTION_ALIGN
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    # apply fill to the merged-away cells too so the banner renders fully if unmerged
    for c in range(2, n_cols + 1):
        cc = ws.cell(row=start_row, column=c)
        cc.fill = section_fill
        cc.font = SECTION_FONT
        cc.alignment = SECTION_ALIGN
    ws.row_dimensions[start_row].height = SECTION_ROW_HEIGHT
    start_row += 1

    # column header
    for c, name in enumerate(header_cols, start=1):
        cell = ws.cell(row=start_row, column=c, value=name)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
    start_row += 1

    # data rows
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=start_row, column=c, value=val)
            cell.alignment = WRAP
        start_row += 1
    return start_row


def build_tab(
    wb: Workbook,
    name: str,
    train_title: str,
    bench_title: str,
    gen_rows: list[list],
    bench_rows: list[list],
) -> int:
    ws = wb.create_sheet(title=name)
    # set column widths (use widest layout)
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    next_row = 1
    next_row = write_section(
        ws, train_title, TRAIN_HEADER_COLS, gen_rows, next_row, TRAIN_SECTION_FILL
    )
    # blank spacer
    next_row += 1
    next_row = write_section(
        ws, bench_title, BENCH_HEADER_COLS, bench_rows, next_row, BENCH_SECTION_FILL
    )

    # freeze the very top
    ws.freeze_panes = "A2"
    return next_row - 1  # last used row


def main() -> None:
    assert GEN_DIR.exists(), f"Missing gen dir: {GEN_DIR}"
    assert BENCH_DIR.exists(), f"Missing bench dir: {BENCH_DIR}"

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    summary: list[tuple[str, int, int, int]] = []

    for bench_name, gen_file, gen_letter, train_title, bench_title, bench_file in BENCHMARKS:
        gen_path = GEN_DIR / gen_file
        bench_path = BENCH_DIR / bench_file
        if not gen_path.exists():
            print(f"[skip] missing gen file: {gen_path}")
            continue
        if not bench_path.exists():
            print(f"[skip] missing bench file: {bench_path}")
            continue

        gen_items = read_jsonl(gen_path, limit=SAMPLE_N)

        # Pull only WRONG predictions (correct == False), up to SAMPLE_N.
        all_bench = read_jsonl(bench_path, limit=None)
        wrong_only = [b for b in all_bench if b.get("correct") is False][:SAMPLE_N]

        gen_rows = [gen_sample_to_row(i + 1, c) for i, c in enumerate(gen_items)]
        bench_rows = [benchmark_sample_to_row(i + 1, b) for i, b in enumerate(wrong_only)]

        last = build_tab(wb, bench_name, train_title, bench_title, gen_rows, bench_rows)
        summary.append((bench_name, len(gen_rows), len(bench_rows), last))
        print(
            f"[ok] {bench_name}: gen={len(gen_rows)} "
            f"bench_wrong={len(bench_rows)}/{len(all_bench)} total_rows={last}"
        )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print("\nTab summary:")
    for name, ng, nb, last in summary:
        print(f"  {name:14s} training={ng:3d}  benchmark_wrong={nb:3d}  last_row={last}")


if __name__ == "__main__":
    main()
